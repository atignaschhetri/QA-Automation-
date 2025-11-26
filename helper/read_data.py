from openpyxl import load_workbook

def read_all_data(path, sheet):
        wb = load_workbook(path)
        ws = wb[sheet]

        data_list = []
        # Read each row starting from row 2
        for row in ws.iter_rows(min_row=2, values_only=True):
            data_list.append(list(row))
        return data_list


if __name__ == "__main__":
    data = read_all_data(r"C:\Users\Hp\PycharmProjects\PythonProject\automations\Book1.xlsx", "Sheet1")

    print(data)